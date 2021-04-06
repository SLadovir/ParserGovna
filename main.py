from openpyxl import load_workbook


def is_number(str):
    if str:
        return str.isdigit()
    else:
        return False
    # try:
    #     float(str)
    #     return True
    # except ValueError:
    #     return False


def findShitInH(wsg):

    sortOfShit = []

def findPodvidGovn(wsg):
    # тест коллекции говна
    # sortOfGovn.append(wsg['F6'].value)
    # print(sortOfGovn.count(wsg['F6']))
    # for govno in sortOfGovn:
    #     # if sortOfGovn.count(wsg['F6'].value) == 0:
    #     print(govno)

    sortOfGovn = []  # коллекция говна

    for i in range(3, 11361):  # Парсинг корпусов (говна) 1(этап)
        if wsg['F' + i.__str__()].value and (sortOfGovn.count(wsg['F' + i.__str__()].value) == 0):
            sortOfGovn.append(wsg['F' + i.__str__()].value)
    # for govno in sortOfGovn: # вывод разновидностей говна
    # print(govno)

    sortOfGovn2 = []  # коллекция говна для 2-го этапа (парсинг по ;)
    for govno in sortOfGovn:
        sortOfGovn2.append(govno.replace(' ', '').split(';'))
    # for govno2 in sortOfGovn2:
    #     print(govno2)

    sortOfGovn3 = []  # коллекция говна для 2-го этапа (парсинг по -)
    for govno2 in sortOfGovn2:
        for i in govno2:
            sortOfGovn3.append(i.replace('—', '-').replace('–', '-').split(
                '-'))  # Заменяем длинные минусы на короткие и разделяем строки по минусам
    # for govno3 in sortOfGovn3:
    #     print(govno3)

    sortOfGovn4 = []  # подготовка для создания коллекции подвидов говна
    for govno3 in sortOfGovn3:
        if len(govno3) > 1:
            sortOfGovn4.append(govno3[1].split(','))
    # for govno4 in sortOfGovn4:
    #     print(govno4)

    podvidGovn = []  # коллекция подвидов говна

    for govno in sortOfGovn4:  # Получение всех подвидов говна (корпусов, которые могут улететь нахуй)
        for i in govno:
            if podvidGovn.count(i) == 0:
                podvidGovn.append(i)
    podvidGovn.sort()  # сортируем подвид говна

    mainGovn = []  # меин говно

    for govn in sortOfGovn3:
        if mainGovn.count(govn[0])==0:
            mainGovn.append(govn[0])
    # for g in mainGovn:
    #     print(g)

    return mainGovn, podvidGovn

def moveMainGovno(wsg,mainGovn):
    # нужно организовать поиск этого говна в первом столбце
    startX = []  # строки начала Херни

    # print(is_number(1))

    for i in range(3, 11361):  # парсим по первому столбцу
        # print(wsg['A' + i.__str__()].value)
        # print(is_number(wsg['A' + i.__str__()].value))
        if is_number(wsg['A' + i.__str__()].value) == True:  # Ищем порядковый номер Херни
            startX.append(i)
    for i in startX:
        print(i)


        # и добавляем его
            # print(wsg['A' + i.__str__()])
            #
    # for i in startX:
    #     print(i)
            # и после него считаем количество строк до следующего порядкового номера
            # sortOfGovn.append(wsg['F' + i.__str__()].value)
    print(startX)







if __name__ == '__main__':
    wbg = load_workbook('./govno.xlsx')
    wsg = wbg.active

    mainGovn, podvidGovn = findPodvidGovn(wsg)  # получение подвидов говна
    # for g in mainGovn:
    #     print(g)
    # for i in podvidGovn:
    #     print(i)

    moveMainGovno(wsg, mainGovn) # убираем говно (начнем с мейн говна, то есть, если есть мейн
    #  говно, то мы убираем все, что идет после мейн говна в нужную ячейку)
























    wbg.save("govno2.xlsx")  # сохраняем в новую табличку












    # просто пусть будет
    # govna4 = govna3[1].split(',') #Получаем подвиды говна (которые через запятую) podvidGovna
    # print(govna4)
    # print('***************')
    # govna5Final = []

    # for govn in govna4: # склейка главного говна с подвидами (получение различных видов говна)
    #     govna5Final.append(govna3[0]+'-'+govn)
    # print(govna5Final)

    # for govno in sortOfGovn:
    #     if sortOfGovn2.count(wsg['F' + i.__str__()].value) == 0:
    #         sortOfGovn.append(wsg['F' + i.__str__()].value)

    # for govno in sortOfGovn:
    # if wsg['F' + i.__str__()].find(govno):
    #     sortOfGovn.append(wsg['F' + i.__str__()])

    # for i in range(3, 12000): # изменение говна
    #     wsg['A' + i.__str__()] = 42

    # wsg.append([1, 2, 3])
    # import datetime
    # wsg['A2'] = datetime.datetime.now()

    # wb = Workbook() # Создать новую таблицу
